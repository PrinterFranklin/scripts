import openpyxl
import requests
import sys

CVE_COLUMN = 3
API_URL = "https://access.redhat.com/hydra/rest/securitydata"
INPUT_FILE = 'input.xlsx'
OUTPUT_FILE = 'output.xlsx'

def get_data(query):

    full_query = API_URL + query
    r = requests.get(full_query)

    # if r.status_code != 200:
    #     print('ERROR: Invalid request; returned {} for the following '
    #           'query: \n{}'.format(r.status_code, full_query))
    #     sys.exit(1)
    
    # if not r.json():
    #     print('No data returned with the following query:')
    #     print(full_query)
    #     sys.exit(0)

    return r.json()


# open the source workbook
src_wb = openpyxl.load_workbook(filename=INPUT_FILE)
src_sheet = src_wb.active

# get CVEs from column 3
# use set to remove duplicate CVEs
cves = set()
for r in range(2,src_sheet.max_row+1):
    cell_value = src_sheet.cell(row=r, column=CVE_COLUMN).value
    s = cell_value.split(',')
    for cve in s:
        cve = cve.strip('\'[] ')
        cves.add(cve)

cves = list(cves)
cves.sort(reverse=True)

# setup a new workbook
rst_wb = openpyxl.Workbook()
rst_sheet = rst_wb.active

# initialize sheet header
header = ['漏洞编号', '漏洞风险级别', '漏洞描述', '漏洞扫描工具', '修复方式(补丁修复或rebase修复)', '修复版本']
rst_sheet.append(header)

for cve in cves:
    entry = []
    entry.append(cve)
    print(cve)
    try:
        data = get_data('/cve/' + cve + '.json')
        if isinstance(data, dict):
            entry.append(data['threat_severity'])
            entry.append(data['details'][0])
    except requests.exceptions.SSLError:
        print("SSLError occurs when fetching ", cve)
    except KeyError:
        print("KeyError occurs when fetching ", cve)
    print(entry)
    rst_sheet.append(entry)

# save result workbook
print('Saving to result.xlsx...')
rst_wb.save(OUTPUT_FILE)
print('Done')
